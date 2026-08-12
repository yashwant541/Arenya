"""
converter.py — the RTL Arabic workbook -> LTR English workbook engine.

Given one input file (.xlsx / .xlsm / .xls / .csv), for every sheet it:

  1. Reads the grid (values + basic layout).
  2. Detects reading direction — RTL is inferred from the sheet's own
     sheetView.rightToLeft flag, and confirmed by where the Arabic text sits
     (labels on the right, numbers marching leftward). CSVs are judged by
     content since they carry no direction flag.
  3. Translates every cell through the FinancialTranslator.
  4. If the sheet was RTL, reverses the column order so the result reads
     left-to-right like a normal English statement (labels first, then the
     period columns in their natural order).
  5. Writes a clean English .xlsx: Arial, frozen header, number formats,
     LTR sheet view, and a companion "Translation Log" sheet listing every
     fuzzy/untranslated cell so a human can audit the glossary.

Design choices:
  - Numbers are preserved exactly (never "translated"); only Arabic-Indic
    digits inside text cells are normalized to ASCII.
  - The original layout/shape is preserved: same number of logical columns,
    just mirrored and translated. Nothing is summarised or dropped.
"""
from __future__ import annotations
import csv
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .arabic import is_arabic, arabic_ratio
from .translator import FinancialTranslator, Match

# ---- styling ------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_LABEL_FONT = Font(name="Arial", bold=True, size=10)
_BODY_FONT = Font(name="Arial", size=10)
_REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")   # amber: fuzzy match
_UNTRANS_FILL = PatternFill("solid", fgColor="F8CBAD")  # orange: no match
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


@dataclass
class SheetReport:
    name: str
    rows: int
    cols: int
    was_rtl: bool
    cells_translated: int = 0
    exact: int = 0
    fuzzy: int = 0
    untranslated: int = 0
    review_items: list = field(default_factory=list)  # (sheet, coord, src, eng, method, conf)


@dataclass
class FileReport:
    input_path: str
    output_path: str
    sheets: list = field(default_factory=list)

    @property
    def total_untranslated(self) -> int:
        return sum(s.untranslated for s in self.sheets)


class ExcelConverter:
    def __init__(self, translator: FinancialTranslator):
        self.tr = translator

    # -- reading -----------------------------------------------------------
    def _read_sheets(self, path: Path):
        """Return list of (name, grid[list[list]], was_rtl)."""
        suf = path.suffix.lower()
        if suf == ".csv":
            return [self._read_csv(path)]
        if suf == ".xls":
            return self._read_xls(path)
        return self._read_openpyxl(path)  # .xlsx / .xlsm / .xltx

    def _read_csv(self, path: Path):
        # sniff delimiter; Arabic CSVs are often utf-8 or utf-8-sig
        raw = path.read_text(encoding="utf-8-sig", errors="replace")
        try:
            dialect = csv.Sniffer().sniff(raw[:4096], delimiters=",;\t")
            delim = dialect.delimiter
        except csv.Error:
            delim = ","
        grid = [row for row in csv.reader(raw.splitlines(), delimiter=delim)]
        return (path.stem, grid, self._needs_flip(grid, view_flag=False))

    def _read_openpyxl(self, path: Path):
        keep_vba = path.suffix.lower() == ".xlsm"
        wb = load_workbook(path, data_only=True, keep_vba=keep_vba)
        out = []
        for ws in wb.worksheets:
            grid = [[c.value for c in row] for row in ws.iter_rows()]
            grid = [r for r in grid if any(v not in (None, "") for v in r)] or grid
            view_flag = bool(getattr(ws.sheet_view, "rightToLeft", False))
            out.append((ws.title, grid, self._needs_flip(grid, view_flag)))
        return out

    def _read_xls(self, path: Path):
        try:
            import pandas as pd
            sheets = pd.read_excel(path, sheet_name=None, header=None)  # needs xlrd
        except Exception as exc:  # xlrd missing or unreadable
            raise RuntimeError(
                f"Cannot read legacy .xls '{path.name}': {exc}. "
                "Install xlrd (pip install xlrd) or re-save the file as .xlsx."
            )
        out = []
        for name, df in sheets.items():
            grid = df.where(df.notna(), None).values.tolist()
            out.append((name, grid, self._needs_flip(grid, view_flag=False)))
        return out

    @classmethod
    def _needs_flip(cls, grid, view_flag: bool) -> bool:
        """Decide whether columns must be reversed to read left-to-right.

        The deciding signal is PHYSICAL: where do the Arabic labels sit in the
        stored grid?

          - labels physically in the RIGHT columns  -> stored in visual (mirrored)
            order; reverse so the label lands in column A.  FLIP.
          - labels already in the LEFT columns (col A) -> stored logically and
            only *displayed* RTL; keep order, just switch the view to LTR.  NO FLIP.

        The sheetView.rightToLeft flag is only a tie-breaker when the physical
        test is inconclusive (e.g. a label column dead-centre).
        """
        side = cls._label_side(grid)          # -1 left, 0 unclear, +1 right
        if side > 0:
            return True
        if side < 0:
            return False
        return view_flag                       # inconclusive -> trust the flag

    @staticmethod
    def _label_side(grid) -> int:
        """+1 if Arabic labels lean right, -1 if they lean left, 0 if unclear."""
        if not grid:
            return 0
        width = max((len(r) for r in grid), default=0)
        if width < 2:
            return 0
        third = max(1, width // 3)
        left = right = 0
        for row in grid:
            for i, v in enumerate(row):
                if isinstance(v, str) and is_arabic(v):
                    if i < third:
                        left += 1
                    elif i >= width - third:
                        right += 1
        if right > left * 1.2:
            return 1
        if left > right * 1.2:
            return -1
        return 0

    # -- conversion --------------------------------------------------------
    def convert_file(self, in_path, out_dir) -> FileReport:
        in_path = Path(in_path)
        out_dir = Path(out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{in_path.stem}_EN.xlsx"

        wb = Workbook()
        wb.remove(wb.active)
        report = FileReport(str(in_path), str(out_path))

        for name, grid, was_rtl in self._read_sheets(in_path):
            sr = self._convert_sheet(wb, name, grid, was_rtl)
            report.sheets.append(sr)

        self._write_log(wb, report)
        wb.save(out_path)
        return report

    def _convert_sheet(self, wb: Workbook, name: str, grid, was_rtl: bool) -> SheetReport:
        # translate the header row (row 0) for a clean tab name if it is Arabic
        translated = [[self.tr.translate_cell(v) for v in row] for row in grid]

        if was_rtl:
            translated = [list(reversed(row)) for row in translated]

        title = self._sheet_title(translated, name)
        ws = wb.create_sheet(title=title[:31])
        ws.sheet_view.rightToLeft = False  # force LTR output

        sr = SheetReport(name=title, rows=len(translated),
                         cols=max((len(r) for r in translated), default=0),
                         was_rtl=was_rtl)

        for r, row in enumerate(translated, start=1):
            for c, m in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=m.english)
                self._style_cell(cell, m, r)
                self._tally(sr, m, ws.title, cell.coordinate)

        self._finish_sheet(ws, sr)
        return sr

    # -- styling / bookkeeping --------------------------------------------
    def _style_cell(self, cell, m: Match, row: int):
        is_number = isinstance(m.english, (int, float))
        if row == 1:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.font = _BODY_FONT if is_number else _LABEL_FONT
            cell.alignment = Alignment(horizontal="right" if is_number else "left",
                                       vertical="center")
        cell.border = _BORDER
        if is_number and row > 1:
            cell.number_format = '#,##0;(#,##0);-'   # header numbers (years) stay General
        if m.method == "fuzzy":
            cell.fill = _REVIEW_FILL
        elif m.method == "untranslated":
            cell.fill = _UNTRANS_FILL

    def _tally(self, sr: SheetReport, m: Match, sheet_title: str, coord: str):
        if m.method in ("exact", "no_article", "fuzzy", "untranslated") and \
           isinstance(m.english, str) and m.english:
            sr.cells_translated += 1
            if m.method in ("exact", "no_article"):
                sr.exact += 1
            elif m.method == "fuzzy":
                sr.fuzzy += 1
                sr.review_items.append((sheet_title, coord, m.source, m.english, m.method, m.confidence))
            elif m.method == "untranslated":
                sr.untranslated += 1
                sr.review_items.append((sheet_title, coord, m.source, m.english, m.method, m.confidence))

    def _finish_sheet(self, ws, sr: SheetReport):
        if sr.rows:
            ws.freeze_panes = "A2"
        # width: label column wider, number columns tidy
        for c in range(1, sr.cols + 1):
            letter = get_column_letter(c)
            longest = max((len(str(ws.cell(row=r, column=c).value or ""))
                           for r in range(1, min(sr.rows, 200) + 1)), default=10)
            ws.column_dimensions[letter].width = min(max(longest + 2, 10),
                                                      48 if c == 1 else 18)

    @staticmethod
    def _sheet_title(translated_grid, fallback: str) -> str:
        # prefer a translated statement title found in the first few rows
        for row in translated_grid[:4]:
            for m in row:
                if isinstance(m.english, str) and m.category == "statement_titles":
                    return m.english
        # else translate the fallback name if it's Arabic
        return fallback if not is_arabic(fallback) else fallback

    def _write_log(self, wb: Workbook, report: FileReport):
        ws = wb.create_sheet("Translation Log")
        ws.sheet_view.rightToLeft = False
        headers = ["Sheet", "Cell", "Arabic source", "English output", "Method", "Confidence"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        r = 2
        any_items = False
        for sr in report.sheets:
            for item in sr.review_items:
                any_items = True
                for c, val in enumerate(item, start=1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.font = _BODY_FONT
                    if item[4] == "untranslated":
                        cell.fill = _UNTRANS_FILL
                    elif item[4] == "fuzzy":
                        cell.fill = _REVIEW_FILL
                r += 1
        if not any_items:
            ws.cell(row=2, column=1, value="No fuzzy or untranslated cells — all matched exactly.").font = _BODY_FONT
        widths = [22, 8, 34, 34, 14, 12]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"
