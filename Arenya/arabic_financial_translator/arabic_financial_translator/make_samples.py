"""Create realistic Arabic RTL sample spreadsheets to exercise the converter."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import csv

OUT = Path(__file__).parent / "samples"
OUT.mkdir(exist_ok=True)
AR = Font(name="Arial", size=11)


def _arabic_view(ws):
    ws.sheet_view.rightToLeft = True
    for row in ws.iter_rows():
        for c in row:
            c.font = AR
            c.alignment = Alignment(horizontal="right")


# ---------------------------------------------------------------------------
# Sample 1: Balance sheet stored in VISUAL (mirrored) order.
# Physical column layout (A..D):  2024 | 2025 | Note | Label
# i.e. the label sits in the RIGHTMOST stored column -> converter must FLIP.
# ---------------------------------------------------------------------------
def balance_sheet_visual():
    wb = Workbook()
    ws = wb.active
    ws.title = "الميزانية"
    rows = [
        ["٢٠٢٤", "٢٠٢٥", "إيضاح", "قائمة المركز المالي الموحدة"],
        [None, None, None, "الأصول"],
        [63447, 77746, "13", "النقد والأرصدة لدى البنوك المركزية"],
        [43593, 43901, "15", "القروض والسلف للبنوك"],
        [281032, 286788, "15", "القروض والسلف للعملاء"],
        [144556, 166956, None, "الاستثمارات في الأوراق المالية"],
        [81472, 65782, "14", "الأدوات المالية المشتقة"],
        [5791, 6231, "17", "الشهرة والأصول غير الملموسة"],
        [43468, 67931, None, "أصول أخرى"],
        [849688, 919955, None, "إجمالي الأصول"],
        [None, None, None, "المطلوبات"],
        [25400, 30846, None, "ودائع البنوك"],
        [464489, 530161, None, "حسابات العملاء"],
        [64609, 72858, "22", "سندات الدين المصدرة"],
        [10382, 8834, None, "المطلوبات الثانوية"],
        [798404, 865369, None, "إجمالي المطلوبات"],
        [None, None, None, "حقوق الملكية"],
        [6695, 6614, None, "رأس المال وعلاوة الإصدار"],
        [8724, 10406, None, "احتياطيات أخرى"],
        [28969, 29573, None, "الأرباح المبقاة"],
        [51284, 54586, None, "إجمالي حقوق الملكية"],
        [849688, 919955, None, "إجمالي حقوق الملكية والمطلوبات"],
    ]
    for r in rows:
        ws.append(r)
    _arabic_view(ws)
    wb.save(OUT / "balance_sheet_ar.xlsx")


# ---------------------------------------------------------------------------
# Sample 2: Income statement stored in LOGICAL order but displayed RTL.
# Physical layout (A..D):  Label | Note | 2025 | 2024
# label is already in column A -> converter must NOT flip, only set LTR view.
# ---------------------------------------------------------------------------
def income_statement_logical():
    wb = Workbook()
    ws = wb.active
    ws.title = "قائمة الدخل"
    rows = [
        ["قائمة الدخل الموحدة", "إيضاح", "٢٠٢٥", "٢٠٢٤"],
        ["إيرادات الفوائد", None, 24547, 27862],
        ["مصروفات الفوائد", None, -18592, -21496],
        ["صافي إيرادات الفوائد", "3", 5955, 6366],
        ["صافي إيرادات الرسوم والعمولات", "4", 4249, 3734],
        ["صافي إيرادات المتاجرة", "5", 10294, 9615],
        ["إيرادات تشغيلية أخرى", "6", 444, -172],
        ["إجمالي الإيرادات التشغيلية", None, 20942, 19543],
        ["تكاليف الموظفين", None, -9109, -8510],
        ["المصروفات العمومية والإدارية", None, -2591, -2465],
        ["الاستهلاك والإطفاء", None, -1170, -1126],
        ["إجمالي المصروفات التشغيلية", "7", -13304, -12502],
        ["مخصص انخفاض قيمة الائتمان", "8", -672, -547],
        ["الربح قبل الضريبة", None, 6963, 6014],
        ["الضريبة", "10", -1866, -1972],
        ["ربح السنة", None, 5097, 4042],
        ["الحقوق غير المسيطرة", None, 12, -8],
        ["ربحية السهم الأساسية", "12", 195.4, 141.3],
    ]
    for r in rows:
        ws.append(r)
    _arabic_view(ws)
    wb.save(OUT / "income_statement_ar.xlsx")


# ---------------------------------------------------------------------------
# Sample 3: CSV in visual/mirrored order (no direction flag exists in CSV),
# with a deliberately misspelled label to exercise the FUZZY path.
# ---------------------------------------------------------------------------
def cashflow_csv():
    rows = [
        ["٢٠٢٤", "٢٠٢٥", "قائمة التدفقات النقدية الموحدة"],
        [7041, 7638, "الربح قبل الضريبه"],                       # teh marbuta variant
        [1126, 1170, "الاستهلاك والاطفاء"],
        [8804, 9200, "صافي النقد من الانشطة التشغيليه"],        # spelling drift -> fuzzy
        [-3200, -4100, "شراء الممتلكات والمعدات"],
        [63447, 77746, "النقد وما في حكمه في نهاية السنة"],
        [500, 620, "بند غير معروف تماماً"],                     # unknown -> untranslated
    ]
    with open(OUT / "cash_flow_ar.csv", "w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows(rows)


if __name__ == "__main__":
    balance_sheet_visual()
    income_statement_logical()
    cashflow_csv()
    print("Wrote samples:", *[p.name for p in sorted(OUT.iterdir())])
